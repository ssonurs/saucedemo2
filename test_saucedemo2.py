from datetime import date
from selenium import webdriver
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.by import By
from pathlib import Path
from constants import *
import pytest
import openpyxl

class Test_Odev2:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(BASE_DOMAIN_URL)
    def teardown_method(self, method):
        self.driver.quit()

    # locked_out_user ile giriş yapıldığında verilen uyarı mesajının doğrulanması.
    @pytest.mark.parametrize("username1,password1",[(LOCKED_OUT_USER_LOGIN,PASSWORD)])
    def test_loginWithLockedID(self,username1,password1):
        usernameID = USERNAME_ID
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,usernameID)))
        username = self.driver.find_element(By.ID, usernameID)
        username.send_keys(username1)

        passwordID = PASSWORD_ID
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,passwordID)))
        password = self.driver.find_element(By.ID, passwordID)
        password.send_keys(password1)

        loginBtn = self.driver.find_element(By.ID, LOGIN_BTN_ID)
        loginBtn.click()
        
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.CLASS_NAME,ERROR_MESSAGE_CLASS_NAME)))
        errorMessage = self.driver.find_element(By.CLASS_NAME, ERROR_MESSAGE_CLASS_NAME)
        assert errorMessage.text == ERROR_MESSAGE_TEXT
        
    @pytest.mark.parametrize("username1,password1",[(STANDARD_USER,PASSWORD)])
    def test_login(self,username1,password1):
        usernameID = USERNAME_ID
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,usernameID)))
        username = self.driver.find_element(By.ID, usernameID)
        username.send_keys(username1)

        passwordID = PASSWORD_ID
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,passwordID)))
        password = self.driver.find_element(By.ID, passwordID)
        password.send_keys(password1)

        loginBtn = self.driver.find_element(By.ID, LOGIN_BTN_ID)
        loginBtn.click()
    
    #normal giriş yapılıp sonrasında ürünlerin sayısının doğrulanması
    def test_numberOfProduct(self):

        self.test_login(STANDARD_USER,PASSWORD)
        productList = []

        products = self.driver.find_elements(By.ID,PRODUCTS_ID)
        lenProducts = len(products)

        for i in range(lenProducts):
            product = products[i]
            productList.append(product)

        assert len(productList) == lenProducts    

    # excelden ürünleri okuma 
    def readFromExcelFile():
        excelFile = openpyxl.load_workbook(PRODUCTS_NAME_EXCEL)
        selectedSheet = excelFile[PAGE_1]

        rows = selectedSheet.max_row

        data = []

        for i in range(2,rows+1):
            productName = selectedSheet.cell(i,1).value
            data.append(productName)
        return data 

    # ürünlerin isimlerinin excel dosyalarındaki isimlerle uyuşması
    def test_namesAreEqual(self):
        
        excelFile = openpyxl.load_workbook(PRODUCTS_NAME_EXCEL)
        selectedSheet = excelFile[PAGE_1]

        rows = selectedSheet.max_row

        data = []

        for i in range(2,rows+1):
            productName = selectedSheet.cell(i,1).value
            data.append(productName)

        self.test_login(STANDARD_USER,PASSWORD)
        productNames = self.driver.find_elements(By.CLASS_NAME,PRODUCT_NAMES_CLASS_NAME)
        names = []
        
        for i in range(len(productNames)):
            product = productNames[i]
            productText = product.text
            names.append(productText)
        
        assert names[0] == data[0]    

    # ürünlerin z'den a ya sıralanma fonksiyonun test edilmesi
    def test_zToA(self):
        self.test_login(STANDARD_USER,PASSWORD)    
        filterBtnZtoA = self.driver.find_element(By.XPATH, FILTER_BTN_ZTOA_XPATH)
        filterBtnZtoA.click()

        products = self.driver.find_elements(By.CLASS_NAME,PRODUCT_NAMES_CLASS_NAME)
        productList = []

        for i in range(0,len(products)):
            product = products[i]
            productText = product.text
            productList.append(productText)

        
        reverseList = sorted(productList,reverse=True)
        assert reverseList[0] == TEST_TISORT

        
    # ürünlerin düşük fiyattan yüksek fiyata sıralanma fonksiyonun test edilmesi
    def test_priceLowToHigh(self):
        self.test_login(STANDARD_USER,PASSWORD)
        
        filterBtnLowToHigh = self.driver.find_element(By.XPATH,FILTER_BTN_LOWTOHIGH_XPATH)
        filterBtnLowToHigh.click()

        productPrices = self.driver.find_elements(By.CLASS_NAME ,PRODUCT_PRICES_CLASS_NAME)

        priceList = []

        for i in range(0,len(productPrices)):
            price = productPrices[i]
            priceText = float(price.text.replace("$",""))
            priceList.append(priceText)

        sortedList = sorted(priceList)
        assert sortedList[0] == 7.99  

    # bir excel dosyasında ismi geçen ürünlerin sepete eklenmesi fonksiyonu testi
    def test_excelBasket(self):
        self.test_login(STANDARD_USER,PASSWORD)

        excelFile = openpyxl.load_workbook(BELIRLI_URUNLER_EXCEL)
        selectedSheet = excelFile[PAGE_1]

        rows = selectedSheet.max_row
    
        data = []

        for i in range(2,rows+1):
            productName = selectedSheet.cell(i,1).value
            data.append(productName)

        
        if productName == PRODUCT_NAME1: 
            addChart = self.driver.find_element(By.ID,ADD_TO_CARD_PRODUCTNAME1)
            addChart.click()
        
            basket = self.driver.find_elements(By.CLASS_NAME,BASKET_CLASS_NAME)
            sizeBasket = len(basket)
            assert sizeBasket > 0

        elif productName == PRODUCT_NAME2: 
            addChart = self.driver.find_element(By.ID,ADD_TO_CARD_PRODUCTNAME2)
            addChart.click()
        
            basket = self.driver.find_elements(By.CLASS_NAME,BASKET_CLASS_NAME)
            sizeBasket = len(basket)
            assert sizeBasket > 0


    #sepete eklenen ürünlerin sepet sayfasında doğru bir şekilde görünmesi testi
    def test_productBasket(self):
        self.test_login(STANDARD_USER,PASSWORD)

        addToCard = self.driver.find_element(By.ID,ADD_TO_CARD_ID)
        addToCard.click()
        
        basket = self.driver.find_element(By.CLASS_NAME,BASKET_CLASS_NAME)
        basket.click()
        
        nameField = self.driver.find_element(By.CLASS_NAME,PRODUCT_NAMES_CLASS_NAME)
        assert nameField.text == SAUCE_LABS_BACKPACK_TEXT


    #sepetten kaldırılan ürünün sepet ekranından kaldırılma testi
    def test_removeFromBasket(self):
        self.test_login(STANDARD_USER,PASSWORD)

        addToCard = self.driver.find_element(By.ID,ADD_TO_CARD_ID)
        addToCard.click()
        
        basket = self.driver.find_element(By.CLASS_NAME,BASKET_CLASS_NAME)
        basket.click()

        removeBtn = self.driver.find_element(By.ID,REMOVE_BTN_ID )
        removeBtn.click()

        removedCheck = self.driver.find_elements(By.CLASS_NAME,REMOVE_CHECK_CLASS_NAME)
        assert len(removedCheck) > 0